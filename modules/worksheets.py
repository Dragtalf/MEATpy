import os
import re
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule

def write_worksheet(worksheet_path, headers, rows):
    """
    Helper function to write data to a worksheet.
    """
    with open(worksheet_path, 'w') as file:
        file.write("\t".join(headers) + "\n")
        for row in rows:
            file.write("\t".join(str(item) for item in row) + "\n")

def generate_worksheet1(output, genome_ids, hmm_table_head, hmm_table, hmmscan_result, hmmscan_hits):
    """
    Create the first worksheet with HMM presence, hit numbers, and hits per genome.
    """
    worksheet_path = os.path.join(output, "MEATpy_result_worksheet1.tsv")

    # 🔧 Sanitize keys to avoid whitespace mismatches
    hmmscan_result = {gn_id.strip(): v for gn_id, v in hmmscan_result.items()}
    hmmscan_hits = {gn_id.strip(): v for gn_id, v in hmmscan_hits.items()}

    headers = hmm_table_head[1:12] + [
        f"{gn_id.strip()} {suffix}" for gn_id in sorted(genome_ids)
        for suffix in ("Hmm presence", "Hit numbers", "Hits")
    ]

    rows = []

    for temp in sorted(hmm_table.values()):
        parts = temp.split('\t')
        row = parts[1:12]

        # Extract and sanitize HMMs
        hmms_raw = parts[10]
        hmms = [h.strip().split()[0] for h in re.split(r'[;,]', hmms_raw) if h.strip()]

        for gn_id in sorted(genome_ids):
            gn_id = gn_id.strip()
            hit_num = 0
            hits = []
            hit_found = False  # track if any HMM hit exists for this genome

            genome_hits = hmmscan_result.get(gn_id, {})
            genome_hits_strs = hmmscan_hits.get(gn_id, {})

            for hmm in hmms:
                if hmm in genome_hits:
                    hit_found = True
                    hit_num += genome_hits[hmm]

                    hit_string = genome_hits_strs.get(hmm, "None")
                    if hit_string != "None":
                        hits.extend(h.strip() for h in hit_string.split(','))
                    else:
                        hits.append("None")
                else:
                    hits.append("None")
                    logging.debug(f"🔍 HMM {hmm} not found for genome {gn_id}. Genome HMM keys: {list(genome_hits.keys())}")

            hmm_presence = "Present" if hit_found else "Absent"
            hits_output = ";".join(hits)

            row.extend([hmm_presence, str(hit_num), hits_output])

        rows.append(row)

    write_worksheet(worksheet_path, headers, rows)

def generate_worksheet2(output, genome_ids, hmm_table_head, hmm_table_temp_2, hmm_table, hmmscan_result):
    """
    Create the second worksheet focusing on function presence based on HMM results.
    """
    worksheet_path = os.path.join(output, "MEATpy_result_worksheet2.tsv")

    # 3 columns from worksheet 1: Category, Function, Gene_abbreviation
    headers = [hmm_table_head[i] for i in [1, 2, 4]] + [f"{gn_id} Function presence" for gn_id in sorted(genome_ids)]
    rows = []

    for line_no, line_content in sorted(hmm_table_temp_2.items()):
        fields = line_content.split('\t')
        row = fields[2:5]  # (Category, Pathway, Gene_abbreviation)
        ref = fields[1]  # reference to worksheet1 row(s), e.g. "34" or "34||35"

        if "||" not in ref:
            # Single line reference
            hmms_raw = hmm_table[ref].split('\t')[10]
            hmms = hmms_raw.split(', ') if ', ' in hmms_raw else [hmms_raw]
        else:
            # Multiple references joined by ||
            hmm_set = set()
            for sub_ref in ref.split("||"):
                hmms_raw = hmm_table[sub_ref].split('\t')[10]
                if ', ' in hmms_raw:
                    hmm_set.update(hmms_raw.split(', '))
                else:
                    hmm_set.add(hmms_raw)
            hmms = list(hmm_set)

        for gn_id in sorted(genome_ids):
            hmm_presence = "Present" if any(hmmscan_result.get(gn_id, {}).get(hmm, 0) > 0 for hmm in hmms) else "Absent"
            row.append(hmm_presence)

        rows.append(row)

    write_worksheet(worksheet_path, headers, rows)

def generate_worksheet3(output, genome_ids, kegg_module2name, category_modules, module_result):
    """
    Create the third worksheet with KEGG module presence per genome.
    """
    worksheet_path = os.path.join(output, "MEATpy_result_worksheet3.tsv")
    headers = ["Module ID", "Module", "Module Category"] + [f"{gn_id} Module presence" for gn_id in sorted(genome_ids)]
    rows = []

    # Build module → category mapping once
    module2category = {
        mod: cat for cat, mods in category_modules.items() for mod in set(mods)
    }

    for module in sorted(module_result.keys()):
        module_category = module2category.get(module, "")

        row = [
            module,
            kegg_module2name.get(module, ""),
            module_category
        ]
        row.extend(
            module_result[module].get(gn_id, "Absent") for gn_id in sorted(genome_ids)
        )

        rows.append(row)

    write_worksheet(worksheet_path, headers, rows)

def generate_worksheet4(output, genome_ids, step_presence_by_genome, kegg_module, kegg_module2name, category_modules):
    """
    Create the fourth worksheet with detailed module step presence.
    Each row represents a module step (e.g. M00001+01) with KO IDs and genome-wise presence.
    """
    worksheet_path = os.path.join(output, "MEATpy_result_worksheet4.tsv")
    headers = ["Module step", "Module", "KO id", "Module Category"] + [
        f"{gn_id} Module step presence" for gn_id in sorted(genome_ids)
    ]
    rows = []

    for module_step in sorted(step_presence_by_genome.keys()):
        module_id = module_step.split('+')[0]
        module_name = kegg_module2name.get(module_id, "")

        k_string = kegg_module.get(module_step, ["NA"])[0]
        if k_string == "NA":
            logging.warning(f"Missing KO string for module step: {module_step}")

        category = next(
            (cat for cat, mods in category_modules.items() if module_id in mods),
            ""
        )

        row = [module_step, module_name, k_string, category]
        for gn_id in sorted(genome_ids):
            step_presence = step_presence_by_genome[module_step].get(gn_id, '0')
            row.append("Present" if step_presence == '1' else "Absent")

        rows.append(row)

    write_worksheet(worksheet_path, headers, rows)

def generate_worksheet5(output, genome_ids, dbcan_results, dbcan_hits):
    """
    Create the fifth worksheet for CAZyme (dbCAN2) hits.

    Args:
        output (str): Output directory.
        genome_ids (list): List of genome IDs.
        dbcan_results (dict): {genome → {CAZyme → count}}.
        dbcan_hits (dict): {genome → {CAZyme → hit names}}.
    """
    worksheet_path = os.path.join(output, "MEATpy_result_worksheet5.tsv")

    # Gather all unique CAZyme IDs across all genomes
    all_cazyme_ids = set()
    for genome in dbcan_results:
        all_cazyme_ids.update(dbcan_results[genome].keys())
    for genome in dbcan_hits:
        all_cazyme_ids.update(dbcan_hits[genome].keys())

    headers = (
        ["CAZyme ID"] +
        [f"{gn_id} Hit numbers" for gn_id in sorted(genome_ids)] +
        [f"{gn_id} Hits" for gn_id in sorted(genome_ids)]
    )

    rows = []
    for hmm_id in sorted(all_cazyme_ids):
        row = [hmm_id]
        # Hit numbers
        for gn_id in sorted(genome_ids):
            hit_num = dbcan_results.get(gn_id, {}).get(hmm_id, 0)
            row.append(hit_num)
        # Hit names
        for gn_id in sorted(genome_ids):
            raw_hits = dbcan_hits.get(gn_id, {}).get(hmm_id, [])
            if isinstance(raw_hits, list):
                hits_str = ";".join(raw_hits) if raw_hits else "None"
            else:
                hits_str = raw_hits if raw_hits else "None"
            row.append(hits_str)
        rows.append(row)

    write_worksheet(worksheet_path, headers, rows)

def generate_worksheet6(output, genome_ids, merops_out, merops_out_hits, merops_ids):
    """
    Create the sixth worksheet for MEROPS peptidase hits.

    Args:
        output (str): Output directory.
        genome_ids (list): List of genome IDs.
        merops_out (dict): {genome → {MEROPS ID → count}}.
        merops_out_hits (dict): {genome → {MEROPS ID → [gene names]}}.
        merops_ids (set): All MEROPS IDs detected across genomes.
    """
    worksheet_path = os.path.join(output, "MEATpy_result_worksheet6.tsv")

    # Build headers
    headers = (
        ["MEROPS peptidase ID"] +
        [f"{gn_id} Hit numbers" for gn_id in sorted(genome_ids)] +
        [f"{gn_id} Hits" for gn_id in sorted(genome_ids)]
    )

    rows = []
    for merops_id in sorted(merops_ids):
        #logging.info(f"Processing MEROPS ID: {merops_id}")
        row = [merops_id]
        for gn_id in sorted(genome_ids):
            hit_num = merops_out.get(gn_id, {}).get(merops_id, 0)
            #logging.info(f"Hit number for genome {gn_id}: {hit_num}")
            row.append(hit_num)
        for gn_id in sorted(genome_ids):
            #logging.info(f"Processing genome ID: {gn_id}")
            raw_hits = merops_out_hits.get(gn_id, {}).get(merops_id, [])
            #logging.info(f"Raw hits for genome {gn_id}: {raw_hits}")
            if isinstance(raw_hits, list):
                hits = ";".join(raw_hits)
            else:
                hits = raw_hits if raw_hits else "None"
            row.append(hits)
        rows.append(row)

    write_worksheet(worksheet_path, headers, rows)

def combine_worksheets_to_excel(output_dir):

    sheet_names = [
        "HMMHitNum",
        "FunctionHit",
        "KEGGModuleHit",
        "KEGGModuleStepHit",
        "dbCAN2Hit",
        "MEROPSHit"
    ]

    filenames = [
        "MEATpy_result_worksheet1.tsv",
        "MEATpy_result_worksheet2.tsv",
        "MEATpy_result_worksheet3.tsv",
        "MEATpy_result_worksheet4.tsv",
        "MEATpy_result_worksheet5.tsv",
        "MEATpy_result_worksheet6.tsv"
    ]

    abs_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pres_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_font = Font(bold=True)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, filename in zip(sheet_names, filenames):
        filepath = os.path.join(output_dir, filename)
        if not os.path.exists(filepath):
            print(f"⚠️ Skipping missing file: {filepath}")
            continue

        df = pd.read_csv(filepath, sep="\t")
        ws = wb.create_sheet(title=sheet_name)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = bold_font

        col_count = df.shape[1]
        row_count = df.shape[0]
        for col in range(2, col_count + 1):
            coord_range = f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=row_count + 1, column=col).coordinate}"
            ws.conditional_formatting.add(coord_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("Absent", {ws.cell(row=2, column=col).coordinate}))'], fill=abs_fill))
            ws.conditional_formatting.add(coord_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("Present", {ws.cell(row=2, column=col).coordinate}))'], fill=pres_fill))

    output_excel = os.path.join(output_dir, "MEATpy_result.xlsx")
    wb.save(output_excel)
    logging.info(f"Excel spreadsheet created: {output_excel}")
